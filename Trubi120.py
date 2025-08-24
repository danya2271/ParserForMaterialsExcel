import os
import re
from collections import defaultdict
import openpyxl
from docx import Document
import tkinter as tk
from tkinter import filedialog, ttk, scrolledtext
import win32com.client as win32
from tkinter import messagebox
import concurrent.futures
import pythoncom

# --- КОНФИГУРАЦИЯ ---
NAME_KEYWORDS = ['наименование', 'позиция']
MATERIAL_KEYWORDS = ['материал']
LENGTH_KEYWORDS = ['длин', 'метр']
QUANTITY_KEYWORDS = ['кол', 'шт', 'колич']
CONTINGENCY_PERCENTAGE = 10
FILENAME_FILTER_KEYWORD = 'журнал'
EXCLUDE_KEYWORD = 'лист'
# ------------------------------------

def find_columns_indices(header_row):
    indices = {'name': None, 'material': None, 'length': None, 'quantity': None}
    for i, cell_text in enumerate(header_row):
        if not cell_text: continue
        lower_cell_text = str(cell_text).lower()
        if any(kw in lower_cell_text for kw in NAME_KEYWORDS): indices['name'] = i
        if any(kw in lower_cell_text for kw in MATERIAL_KEYWORDS): indices['material'] = i
        if any(kw in lower_cell_text for kw in LENGTH_KEYWORDS): indices['length'] = i
        if any(kw in lower_cell_text for kw in QUANTITY_KEYWORDS): indices['quantity'] = i
    return indices

def parse_value(value):
    if isinstance(value, (int, float)): return value
    if isinstance(value, str):
        try: return float(value.replace(',', '.').strip())
        except (ValueError, TypeError): return 0
    return 0

def natural_sort_key(s):
    return [int(text) if text.isdigit() else text.lower() for text in re.split('([0-9]+)', str(s))]

def parse_doc_in_thread(file_path):
    """
    Обрабатывает один .doc файл в отдельном потоке, создавая свой экземпляр Word.
    Возвращает кортеж из пути к файлу, словаря с данными и вложенного словаря по трубам 102.
    """
    pythoncom.CoInitialize()
    word_app = None
    file_data = defaultdict(float)
    # Структура: {'Труба 102x3.5': {12000: 5, 6000: 2}}
    pipe_102_data_local = defaultdict(lambda: defaultdict(int))
    error_message = None

    try:
        try:
            word_app = win32.Dispatch("Word.Application")
            word_app.Visible = False
            word_app.DisplayAlerts = 0
            word_app.AutomationSecurity = 3
        except Exception as e:
            raise Exception(f"Не удалось запустить MS Word: {e}")

        doc = None
        try:
            doc = word_app.Documents.Open(
                os.path.abspath(file_path),
                ConfirmConversions=False, ReadOnly=True, AddToRecentFiles=False
            )
            for table in doc.Tables:
                try:
                    header_row = table.Rows(1)
                    header_values = [cell.Range.Text.strip('\r\x07 ').strip() for cell in header_row.Cells]
                    column_indices = find_columns_indices(header_values)

                    if column_indices.get('name') is not None and column_indices.get('quantity') is not None:
                        def com_rows_iterator():
                            for i in range(2, table.Rows.Count + 1):
                                yield [cell.Range.Text.strip('\r\x07 ').strip() for cell in table.Rows(i).Cells]

                        last_material_name = ""
                        name_idx = column_indices['name']
                        for row_data in com_rows_iterator():
                            if not any(v for v in row_data if v and str(v).strip()): continue

                            processed_row_data = list(row_data)
                            if len(processed_row_data) <= name_idx: continue

                            name_cell_value = str(processed_row_data[name_idx]).strip()
                            if name_cell_value:
                                last_material_name = name_cell_value
                            else:
                                processed_row_data[name_idx] = last_material_name

                            name_idx_p, material_idx_p, length_col_idx_p, quantity_hdr_idx_p = (
                                column_indices.get('name'), column_indices.get('material'),
                                column_indices.get('length'), column_indices.get('quantity')
                            )
                            if name_idx_p is None or quantity_hdr_idx_p is None: continue

                            name_content = str(processed_row_data[name_idx_p]).strip() if len(processed_row_data) > name_idx_p else ""
                            if not name_content: continue

                            search_text = name_content
                            if material_idx_p is not None and len(processed_row_data) > material_idx_p:
                                search_text += " " + str(processed_row_data[material_idx_p]).strip()

                            if EXCLUDE_KEYWORD in search_text.lower(): continue

                            quantity = 0
                            is_short_row = "L=" in name_content.upper() and len(processed_row_data) < quantity_hdr_idx_p
                            if is_short_row:
                                if len(processed_row_data) > 1: quantity = parse_value(processed_row_data[-1])
                            elif len(processed_row_data) > quantity_hdr_idx_p:
                                quantity = parse_value(processed_row_data[quantity_hdr_idx_p])

                            if quantity <= 0: continue

                            pipe_profile_pattern = r'(102\s*[xх*]\s*\d+(?:[.,]\d+)?)'
                            pipe_match = re.search(pipe_profile_pattern, search_text, re.IGNORECASE)
                            if pipe_match and ('труб' in search_text.lower() or 'тр.' in search_text.lower()):
                                profile_name = pipe_match.group(1).lower().replace(',', '.').replace(' ', '').replace('*', 'x')
                                spec_key = f"Труба {profile_name}"
                                length_mm = 0
                                l_match = re.search(r'L\s*=\s*(\d+)', search_text, re.IGNORECASE)
                                if l_match:
                                    length_mm = parse_value(l_match.group(1))
                                elif length_col_idx_p is not None and len(processed_row_data) > length_col_idx_p:
                                    length_mm = parse_value(processed_row_data[length_col_idx_p])

                                if length_mm > 0:
                                    pipe_102_data_local[spec_key][int(length_mm)] += int(quantity)
                                    file_data[spec_key] += (length_mm / 1000) * quantity
                                    continue

                            rebar_pattern = r'[АаAa][54]00[СсСc]?.*?(?:диаметр|d|D|⌀|ø)\s*(\d+(?:,\d+)?).*?L\s*=\s*(\d+)'
                            profile_with_l_pattern = r'(\d+(?:,\d+)?(?:\s*[хx]\s*\d+(?:,\d+)?){1,2}).*?L\s*=\s*(\d+)'
                            standard_profile_pattern = r'(\d+(?:,\d+)?(?:\s*[хx]\s*\d+(?:,\d+)?){1,2})'

                            rebar_match = re.search(rebar_pattern, search_text, re.IGNORECASE)
                            if rebar_match:
                                diameter = rebar_match.group(1).replace(',', '.'); length_mm = parse_value(rebar_match.group(2))
                                if length_mm > 0: file_data[f"Арматура d {diameter}"] += (length_mm / 1000) * quantity
                                continue

                            profile_l_match = re.search(profile_with_l_pattern, search_text, re.IGNORECASE)
                            if profile_l_match:
                                profile_name = profile_l_match.group(1).replace(',', '.').replace(' ', ''); length_mm = parse_value(profile_l_match.group(2))
                                if length_mm > 0: file_data[f"Профиль {profile_name}"] += (length_mm / 1000) * quantity
                                continue

                            if length_col_idx_p is not None and len(processed_row_data) > length_col_idx_p:
                                profile_std_match = re.search(standard_profile_pattern, search_text)
                                if profile_std_match:
                                    profile_name = profile_std_match.group(1).replace(',', '.').replace(' ', '')
                                    length_mm = parse_value(processed_row_data[length_col_idx_p])
                                    if length_mm > 0: file_data[f"Профиль {profile_name}"] += (length_mm / 1000) * quantity
                except Exception as e_table:
                    print(f"Пропущена таблица в {os.path.basename(file_path)} из-за ошибки: {e_table}")
                    continue
        except Exception as e_doc:
            error_message = f"Ошибка при обработке DOC: {os.path.basename(file_path)} ({e_doc})"
        finally:
            if doc:
                doc.Saved = True
                doc.Close(SaveChanges=False)
    except Exception as e_main:
        error_message = f"Критическая ошибка в потоке для {os.path.basename(file_path)}: {e_main}"
    finally:
        if word_app:
            word_app.Quit(SaveChanges=False)
        pythoncom.CoUninitialize()

    return file_path, file_data, pipe_102_data_local, error_message

class ParserApp(tk.Tk):
    def __init__(self):
        super().__init__()
        self.title("Счётчик 102-х труб (Глубина поиска 2)")
        self.geometry("800x600")
        self.protocol("WM_DELETE_WINDOW", self.on_closing)
        self.word_app = None
        self.pipe_102_data = defaultdict(lambda: defaultdict(int))
        main_frame = ttk.Frame(self, padding="10")
        main_frame.pack(fill="both", expand=True)
        top_frame = ttk.Frame(main_frame)
        top_frame.pack(fill="x", pady=5)
        self.folder_path = tk.StringVar()
        ttk.Label(top_frame, text="Папка для поиска:").pack(side="left", padx=(0, 10))
        ttk.Entry(top_frame, textvariable=self.folder_path, state="readonly", width=50).pack(side="left", fill="x", expand=True)
        ttk.Button(top_frame, text="Выбрать...", command=self.select_folder).pack(side="left", padx=(10, 0))
        self.run_button = ttk.Button(main_frame, text="Запустить анализ", command=self.run_parser)
        self.run_button.pack(fill="x", pady=10)
        self.results_text = scrolledtext.ScrolledText(main_frame, wrap=tk.WORD, height=20, state="disabled")
        self.results_text.pack(fill="both", expand=True)

    def log(self, message):
        self.results_text.config(state="normal")
        self.results_text.insert(tk.END, message + "\n")
        self.results_text.config(state="disabled")
        self.results_text.see(tk.END)
        self.update_idletasks()

    def select_folder(self):
        path = filedialog.askdirectory(title="Выберите папку для сканирования")
        if path:
            self.folder_path.set(path)
            self.log(f"Выбрана папка: {path}")

    def on_closing(self):
        self.destroy()

    def _process_row(self, row_data, column_indices, file_data, pipe_data):
        name_idx, material_idx, length_col_idx, quantity_hdr_idx = (
            column_indices.get('name'), column_indices.get('material'),
            column_indices.get('length'), column_indices.get('quantity')
        )
        if name_idx is None or quantity_hdr_idx is None: return

        name_content = str(row_data[name_idx]).strip() if len(row_data) > name_idx else ""
        if not name_content: return

        search_text = name_content
        if material_idx is not None and len(row_data) > material_idx:
            search_text += " " + str(row_data[material_idx]).strip()

        if EXCLUDE_KEYWORD in search_text.lower(): return

        quantity = 0
        is_short_row = "L=" in name_content.upper() and len(row_data) < quantity_hdr_idx
        if is_short_row:
            if len(row_data) > 1: quantity = parse_value(row_data[-1])
        elif len(row_data) > quantity_hdr_idx:
            quantity = parse_value(row_data[quantity_hdr_idx])

        if quantity <= 0: return

        # --- ОБНОВЛЕННАЯ ЛОГИКА ДЛЯ ИЗВЛЕЧЕНИЯ ПРОФИЛЯ ТРУБЫ ---
        pipe_profile_pattern = r'(102\s*[xх*]\s*\d+(?:[.,]\d+)?)'
        pipe_match = re.search(pipe_profile_pattern, search_text, re.IGNORECASE)
        if pipe_match and ('труб' in search_text.lower() or 'тр.' in search_text.lower()):
            profile_name = pipe_match.group(1).lower().replace(',', '.').replace(' ', '').replace('*', 'x')
            spec_key = f"Труба {profile_name}"
            length_mm = 0
            l_match = re.search(r'L\s*=\s*(\d+)', search_text, re.IGNORECASE)
            if l_match:
                length_mm = parse_value(l_match.group(1))
            elif length_col_idx is not None and len(row_data) > length_col_idx:
                length_mm = parse_value(row_data[length_col_idx])

            if length_mm > 0:
                pipe_data[spec_key][int(length_mm)] += int(quantity)
                file_data[spec_key] += (length_mm / 1000) * quantity
                return

        rebar_pattern = r'[АаAa][54]00[СсСc]?.*?(?:диаметр|d|D|⌀|ø)\s*(\d+(?:,\d+)?).*?L\s*=\s*(\d+)'
        profile_with_l_pattern = r'(\d+(?:,\d+)?(?:\s*[хx]\s*\d+(?:,\d+)?){1,2}).*?L\s*=\s*(\d+)'
        standard_profile_pattern = r'(\d+(?:,\d+)?(?:\s*[хx]\s*\d+(?:,\d+)?){1,2})'

        rebar_match = re.search(rebar_pattern, search_text, re.IGNORECASE)
        if rebar_match:
            diameter = rebar_match.group(1).replace(',', '.'); length_mm = parse_value(rebar_match.group(2))
            if length_mm > 0: file_data[f"Арматура d {diameter}"] += (length_mm / 1000) * quantity
            return

        profile_l_match = re.search(profile_with_l_pattern, search_text, re.IGNORECASE)
        if profile_l_match:
            profile_name = profile_l_match.group(1).replace(',', '.').replace(' ', ''); length_mm = parse_value(profile_l_match.group(2))
            if length_mm > 0: file_data[f"Профиль {profile_name}"] += (length_mm / 1000) * quantity
            return

        if length_col_idx is not None and len(row_data) > length_col_idx:
            profile_std_match = re.search(standard_profile_pattern, search_text)
            if profile_std_match:
                profile_name = profile_std_match.group(1).replace(',', '.').replace(' ', '')
                length_mm = parse_value(row_data[length_col_idx])
                if length_mm > 0: file_data[f"Профиль {profile_name}"] += (length_mm / 1000) * quantity

    def _process_table_iterator(self, rows_iterator, column_indices, file_data, pipe_data):
        last_material_name = ""
        name_idx = column_indices['name']
        for row_data in rows_iterator:
            if not any(v for v in row_data if v and str(v).strip()): continue
            processed_row_data = list(row_data)
            if len(processed_row_data) <= name_idx: continue
            name_cell_value = str(processed_row_data[name_idx]).strip()
            if name_cell_value:
                last_material_name = name_cell_value
            else:
                processed_row_data[name_idx] = last_material_name
            self._process_row(processed_row_data, column_indices, file_data, pipe_data)

    def run_parser(self):
        start_path = self.folder_path.get()
        if not start_path:
            self.log("Ошибка: Папка не выбрана.")
            return
        self.results_text.config(state="normal"); self.results_text.delete('1.0', tk.END); self.results_text.config(state="disabled")
        self.run_button.config(state="disabled")

        master_pipe_data = {}
        self.pipe_102_data.clear()

        self.log(f"Начинаю поиск файлов c '{FILENAME_FILTER_KEYWORD}' в названии (глубина 2)...")
        try:
            all_files = []
            for item_name in os.listdir(start_path):
                item_path = os.path.join(start_path, item_name)
                if os.path.isfile(item_path):
                    if FILENAME_FILTER_KEYWORD in item_name.lower() and not item_name.startswith('~'):
                        all_files.append(item_path)
                elif os.path.isdir(item_path):
                    for sub_item_name in os.listdir(item_path):
                        sub_item_path = os.path.join(item_path, sub_item_name)
                        if os.path.isfile(sub_item_path):
                            if FILENAME_FILTER_KEYWORD in sub_item_name.lower() and not sub_item_name.startswith('~'):
                                all_files.append(sub_item_path)

            all_files.sort(key=natural_sort_key)

            doc_files = [p for p in all_files if p.lower().endswith('.doc')]
            other_files = [p for p in all_files if not p.lower().endswith('.doc')]

            if doc_files:
                self.log(f"\nНачинаю параллельную обработку {len(doc_files)} файлов .doc...")
                max_workers = min(len(doc_files), (os.cpu_count() or 1) * 2)
                with concurrent.futures.ThreadPoolExecutor(max_workers=max_workers) as executor:
                    future_to_path = {executor.submit(parse_doc_in_thread, path): path for path in doc_files}
                    for future in concurrent.futures.as_completed(future_to_path):
                        path = future_to_path[future]
                        relative_path = os.path.relpath(path, start_path)
                        try:
                            file_path_res, file_specific_data, pipe_data_from_file, error_message = future.result()
                            self.log(f"\n[DOC] Обработка завершена: {relative_path}")
                            if error_message: self.log(f"  > {error_message}")
                            if pipe_data_from_file:
                                master_pipe_data[relative_path] = pipe_data_from_file
                                # Мержим вложенные словари для общего итога
                                for profile, length_counts in pipe_data_from_file.items():
                                    for length, count in length_counts.items():
                                        self.pipe_102_data[profile][length] += count
                        except Exception as exc:
                            self.log(f"\n[DOC] КРИТИЧЕСКАЯ ОШИБКА при обработке файла {relative_path}: {exc}")

            for file_path in other_files:
                relative_path = os.path.relpath(file_path, start_path)
                file_ext = os.path.splitext(file_path)[1].lower()
                self.log(f"\n[{file_ext.upper().replace('.', '')}] Обработка: {relative_path}")
                file_specific_data = defaultdict(float)
                file_specific_pipe_data = defaultdict(lambda: defaultdict(int))

                if file_ext == '.xlsx': self.parse_xlsx(file_path, file_specific_data, file_specific_pipe_data)
                elif file_ext == '.docx': self.parse_docx(file_path, file_specific_data, file_specific_pipe_data)

                if file_specific_pipe_data:
                    master_pipe_data[relative_path] = file_specific_pipe_data
                    # Мержим вложенные словари для общего итога
                    for profile, length_counts in file_specific_pipe_data.items():
                        for length, count in length_counts.items():
                            self.pipe_102_data[profile][length] += count

            # --- ВЫВОД ИНФОРМАЦИИ ПО ТРУБАМ ПО КАЖДОМУ ФАЙЛУ ---
            self.log("\n\n###########################################")
            self.log("--- ИНФОРМАЦИЯ ПО ТРУБАМ (ПО ФАЙЛАМ) ---")
            self.log("###########################################")
            if not master_pipe_data:
                self.log("\nТрубы с диаметром 102 не найдены ни в одном файле.")
            else:
                sorted_files = sorted(master_pipe_data.items(), key=lambda item: natural_sort_key(item[0]))
                for filename, pipe_data in sorted_files:
                    self.log(f"\n===========================================\nФАЙЛ: {filename}")
                    sorted_profiles = sorted(pipe_data.items(), key=lambda item: natural_sort_key(item[0]))
                    for profile, length_counts in sorted_profiles:
                        self.log(f"--- {profile} ---")
                        sorted_lengths = sorted(length_counts.items())
                        profile_total_count = sum(length_counts.values())
                        for length, count in sorted_lengths:
                            length_m = f"{(length / 1000):.3f}".replace('.', ',')
                            self.log(f"  Длина: {length_m} м - Количество: {count} шт.")
                        self.log(f"  ИТОГО ({profile}): {profile_total_count} шт.\n")

            # --- ВЫВОД ОБЩЕЙ СВОДНОЙ ИНФОРМАЦИИ ПО ТРУБАМ ---
            self.log("\n\n###########################################")
            self.log("--- СВОДНАЯ ИНФОРМАЦИЯ ПО ТРУБАМ (ОБЩИЙ ИТОГ) ---")
            self.log("###########################################\n")
            if not self.pipe_102_data:
                self.log("Трубы с диаметром 102 не найдены.")
            else:
                total_pipe_count = 0
                sorted_profiles = sorted(self.pipe_102_data.items(), key=lambda item: natural_sort_key(item[0]))
                for profile, length_counts in sorted_profiles:
                    self.log(f"--- {profile} ---")
                    sorted_lengths = sorted(length_counts.items())
                    profile_total_count = sum(length_counts.values())
                    total_pipe_count += profile_total_count
                    for length, count in sorted_lengths:
                        length_m = f"{(length / 1000):.3f}".replace('.', ',')
                        self.log(f"  Длина: {length_m} м - Количество: {count} шт.")
                    self.log(f"  ИТОГО ({profile}): {profile_total_count} шт.\n")
                self.log(f"ОБЩИЙ ИТОГ по всем трубам 102: {total_pipe_count} шт.")


            self.log("\n\n--- Анализ завершен. ---")
        except Exception as e:
            self.log(f"КРИТИЧЕСКАЯ ОШИБКА: {e}")
            import traceback
            self.log(traceback.format_exc())
        finally:
            self.run_button.config(state="normal")

    def parse_xlsx(self, file_path, file_data, pipe_data):
        try:
            workbook = openpyxl.load_workbook(file_path, data_only=True)
            for sheet in workbook.worksheets:
                for row_idx in range(1, sheet.max_row + 1):
                    header_row_values = [cell.value for cell in sheet[row_idx]]
                    column_indices = find_columns_indices(header_row_values)
                    if column_indices.get('name') is not None and column_indices.get('quantity') is not None:
                        rows_iterator = (
                            [cell.value for cell in sheet[data_row_idx]]
                            for data_row_idx in range(row_idx + 1, sheet.max_row + 1)
                        )
                        self._process_table_iterator(rows_iterator, column_indices, file_data, pipe_data)
                        break
        except Exception as e:
            self.log(f"  > Ошибка при чтении файла XLSX: {os.path.basename(file_path)} ({e})")

    def parse_docx(self, file_path, file_data, pipe_data):
        try:
            document = Document(file_path)
            for table in document.tables:
                header_row_values = [cell.text for cell in table.rows[0].cells]
                column_indices = find_columns_indices(header_row_values)
                if column_indices.get('name') is not None and column_indices.get('quantity') is not None:
                    rows_iterator = (
                        [cell.text for cell in row.cells]
                        for row in table.rows[1:]
                    )
                    self._process_table_iterator(rows_iterator, column_indices, file_data, pipe_data)
        except Exception as e:
            self.log(f"  > Ошибка при чтении файла DOCX: {os.path.basename(file_path)} ({e})")

if __name__ == "__main__":
    try:
        import win32com.client
    except ImportError:
        messagebox.showerror("Отсутствует библиотека", "Для обработки .doc файлов необходима библиотека pywin32.\nУстановите ее: pip install pywin32")
        exit()
    app = ParserApp()
    app.mainloop()
